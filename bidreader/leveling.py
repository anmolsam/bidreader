"""Bid leveling — read multiple subcontractor quotes for the same scope and build
an apples-to-apples comparison: bidders as columns, normalized exclusion/scope-gap
matrix (who excluded what), math flags, per-bidder detail.

This is the bid-day workflow estimators actually live in: not "read one quote"
but "compare five subs side-by-side and find the scope differences."

    from bidreader.leveling import level, to_xlsx
    result = level(["subA.pdf", "subB.pdf", "subC.pdf"])
    to_xlsx(result, "leveling.xlsx")
"""
from __future__ import annotations
import os, re
from .extract import read

_STOP = {"the", "and", "for", "all", "any", "with", "from", "this", "that", "are", "not",
         "per", "of", "to", "in", "on", "at", "by", "or", "a", "an", "is", "be", "as"}


def _sig(s):
    return {w for w in re.findall(r"[a-z0-9]+", (s or "").lower()) if len(w) > 2 and w not in _STOP}


def _cluster(records, text_key):
    """Greedy Jaccard clustering of {bidder, text, ...} records into scope rows.
    Returns [{label, tokens, members:{bidder_idx: record}}]."""
    clusters = []
    for r in records:
        toks = _sig(r[text_key])
        best, bestj = None, 0.0
        for c in clusters:
            inter = len(toks & c["tokens"]); union = len(toks | c["tokens"]) or 1
            j = inter / union
            if j > bestj:
                best, bestj = c, j
        if best is not None and bestj >= 0.45:
            best["members"].setdefault(r["bidder"], r)
            best["tokens"] |= toks
        else:
            clusters.append({"label": r[text_key], "tokens": toks, "members": {r["bidder"]: r}})
    return clusters


def level(paths):
    bidders, all_excl, all_gaps = [], [], []
    for i, p in enumerate(paths):
        d = read(p)
        name = d.get("vendor") or os.path.basename(p).rsplit(".", 1)[0]
        bidders.append({
            "idx": i, "name": name, "trade": d.get("trade"),
            "bid_total": d.get("bid_total"), "currency": d.get("currency") or "USD",
            "line_items": d.get("line_items", []),
            "n_items": len(d.get("line_items", [])),
            "n_math_flagged": d.get("math_flagged", 0),
            "exclusions": d.get("exclusions", []),
            "scope_gaps": d.get("scope_gaps", []),
        })
        for e in d.get("exclusions", []):
            all_excl.append({"bidder": i, "item": e.get("item") or e.get("quote", ""),
                             "quote": e.get("quote", ""), "page": e.get("page")})
        for g in d.get("scope_gaps", []):
            all_gaps.append({"bidder": i, "item": g.get("missing", ""), "why": g.get("why", "")})
    return {
        "bidders": bidders,
        "exclusion_rows": _cluster(all_excl, "item"),
        "gap_rows": _cluster(all_gaps, "item"),
    }


# ── text preview (for CLI) ───────────────────────────────────────────────────
def summary_text(result):
    b = result["bidders"]
    out = ["BID LEVELING — " + " vs ".join(x["name"][:18] for x in b), ""]
    out.append(f"{'':28s}" + "".join(f"{x['name'][:16]:>18s}" for x in b))
    def row(label, vals):
        out.append(f"{label:28s}" + "".join(f"{v:>18s}" for v in vals))
    row("Trade", [str(x["trade"] or "-")[:16] for x in b])
    row("Bid total", [f"${x['bid_total']:,.0f}" if x["bid_total"] else "-" for x in b])
    row("Line items", [str(x["n_items"]) for x in b])
    row("Math flags", [str(x["n_math_flagged"]) for x in b])
    row("# exclusions", [str(len(x["exclusions"])) for x in b])
    out += ["", "SCOPE / EXCLUSION MATRIX  (X = this bidder excluded it):"]
    out.append(f"{'exclusion':28s}" + "".join(f"{x['name'][:16]:>18s}" for x in b))
    for c in result["exclusion_rows"]:
        cells = []
        for x in b:
            m = c["members"].get(x["idx"])
            cells.append((f"X p{m['page']}" if m and m.get("page") else "X") if m else "—")
        out.append(f"{c['label'][:27]:28s}" + "".join(f"{v:>18s}" for v in cells))
    return "\n".join(out)


# ── xlsx export ──────────────────────────────────────────────────────────────
def to_xlsx(result, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    b = result["bidders"]
    wb = Workbook()
    ws = wb.active; ws.title = "Leveling"
    HDR = Font(bold=True, color="FFFFFF"); HDRFILL = PatternFill("solid", fgColor="1F4E79")
    SECT = Font(bold=True, color="FFFFFF"); SECTFILL = PatternFill("solid", fgColor="2E75B6")
    EXCL = PatternFill("solid", fgColor="FCE4E4"); GAPF = PatternFill("solid", fgColor="FFF2CC")
    money = '#,##0.00'; thin = Border(*([Side(style="thin", color="D9D9D9")] * 4))
    ncol = 1 + len(b)

    def put(r, c, v, font=None, fill=None, num=None, align=None):
        cell = ws.cell(row=r, column=c, value=v)
        if font: cell.font = font
        if fill: cell.fill = fill
        if num: cell.number_format = num
        cell.alignment = Alignment(horizontal=align or ("left" if c == 1 else "center"), vertical="center", wrap_text=(c == 1))
        cell.border = thin
        return cell

    def section(r, title):
        put(r, 1, title, SECT, SECTFILL)
        for c in range(2, ncol + 1):
            put(r, c, "", SECT, SECTFILL)

    r = 1
    put(r, 1, "BID LEVELING", Font(bold=True, size=14)); r += 1
    put(r, 1, "Bidder →", HDR, HDRFILL)
    for j, x in enumerate(b):
        put(r, 2 + j, x["name"], HDR, HDRFILL)
    r += 1
    section(r, "SUMMARY"); r += 1
    rows = [("Trade", lambda x: x["trade"] or "-", None),
            ("Bid total", lambda x: x["bid_total"], money),
            ("Line items", lambda x: x["n_items"], None),
            ("Arithmetic flags", lambda x: x["n_math_flagged"], None),
            ("# exclusions", lambda x: len(x["exclusions"]), None)]
    for label, fn, num in rows:
        put(r, 1, label, Font(bold=True))
        for j, x in enumerate(b):
            put(r, 2 + j, fn(x), num=num)
        r += 1
    # lowest bid highlight
    totals = [x["bid_total"] for x in b if isinstance(x["bid_total"], (int, float))]
    if totals:
        low = min(totals)
        put(r, 1, "Lowest base bid", Font(bold=True))
        for j, x in enumerate(b):
            mark = "◀ LOW" if x["bid_total"] == low else ""
            put(r, 2 + j, mark, Font(bold=True, color="C00000"))
        r += 1

    r += 1; section(r, "EXCLUSION MATRIX  (filled = this bidder EXCLUDED it → confirm who carries the cost)"); r += 1
    put(r, 1, "Excluded scope", Font(bold=True))
    for j, x in enumerate(b):
        put(r, 2 + j, x["name"], Font(bold=True))
    r += 1
    for c in result["exclusion_rows"]:
        put(r, 1, c["label"])
        for j, x in enumerate(b):
            m = c["members"].get(x["idx"])
            put(r, 2 + j, (f"EXCL (p{m['page']})" if m and m.get("page") else "EXCL") if m else "", fill=EXCL if m else None)
        r += 1

    r += 1; section(r, "SCOPE-GAP MATRIX  (flagged = likely-missing scope to confirm)"); r += 1
    put(r, 1, "Possible gap", Font(bold=True))
    for j, x in enumerate(b):
        put(r, 2 + j, x["name"], Font(bold=True))
    r += 1
    for c in result["gap_rows"]:
        put(r, 1, c["label"])
        for j, x in enumerate(b):
            m = c["members"].get(x["idx"])
            put(r, 2 + j, "⚠" if m else "", fill=GAPF if m else None)
        r += 1

    ws.column_dimensions["A"].width = 46
    for j in range(len(b)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 20
    ws.freeze_panes = "B3"

    # per-bidder detail sheets
    for x in b:
        title = re.sub(r"[\\/*?:\[\]]", " ", x["name"])[:28] or f"Bidder {x['idx']+1}"
        sh = wb.create_sheet(title=title)
        for c, h in enumerate(["CSI", "Description", "Qty", "Unit", "Unit $", "Amount", "Math", "Source page"], 1):
            cell = sh.cell(row=1, column=c, value=h); cell.font = HDR; cell.fill = HDRFILL
        ri = 2
        for li in x["line_items"]:
            sh.cell(row=ri, column=1, value=li.get("section"))
            sh.cell(row=ri, column=2, value=li.get("description"))
            sh.cell(row=ri, column=3, value=li.get("qty"))
            sh.cell(row=ri, column=4, value=li.get("unit"))
            sh.cell(row=ri, column=5, value=li.get("unit_price")).number_format = money
            sh.cell(row=ri, column=6, value=li.get("amount")).number_format = money
            mc = sh.cell(row=ri, column=7, value=li.get("math_check"))
            if li.get("math_check") == "mismatch":
                mc.fill = PatternFill("solid", fgColor="FFC7CE")
            sh.cell(row=ri, column=8, value=li.get("page"))
            ri += 1
        sh.cell(row=ri + 1, column=1, value="EXCLUSIONS").font = Font(bold=True)
        for e in x["exclusions"]:
            ri += 1
            sh.cell(row=ri + 1, column=2, value=f"(p{e.get('page')}) {e.get('quote','')}")
        sh.column_dimensions["B"].width = 60
    wb.save(out_path)
    return out_path
